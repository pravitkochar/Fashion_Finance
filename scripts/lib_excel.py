"""Shared Excel building blocks for event-study briefs.

Public API:
    autosize(ws, max_w=28)
    write_summary_block(ws, rows)               header-styled key/value rows
    write_test_block(ws, title, headers, rows)  bordered table with header fill
    write_overview_charts(ws, paths, anchor_col="A", start_row=None)
    apply_car_conditional_formatting(ws, first_row, last_row, col_indices)
    write_company_event_table(ws, df, cols, start_row)

Each brief composes its own Overview/per-group sheets using these helpers; this
keeps formatting consistent across deliverables without duplicating code.
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill("solid", fgColor="0B2545")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True)
SECTION_FONT = Font(bold=True, size=12)


def autosize(ws: Worksheet, max_w: int = 28) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        w = 8
        for c in col:
            v = c.value
            if v is None:
                continue
            w = max(w, min(max_w, len(str(v)) + 1))
        ws.column_dimensions[letter].width = w


def write_title(ws: Worksheet, text: str, span: str = "A1:H1") -> None:
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(span)


def write_kv_block(ws: Worksheet, kv: list[tuple[str, object]], start_row: int = 3) -> int:
    """Write a list of (label, value) rows. Returns next free row."""
    r = start_row
    for label, value in kv:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=value)
        r += 1
    return r + 1


def write_test_table(ws: Worksheet, title: str, headers: list[str],
                     rows: Iterable[Iterable], start_row: int) -> int:
    """Write a header-styled section with a title and table. Returns next free row."""
    ws.cell(row=start_row, column=1, value=title).font = SECTION_FONT
    hr = start_row + 1
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    r = hr
    for row_vals in rows:
        r += 1
        for c, v in enumerate(row_vals, start=1):
            if isinstance(v, float):
                v = round(v, 4)
            ws.cell(row=r, column=c, value=v)
    return r + 2


def embed_images(ws: Worksheet, paths: Iterable[Path | str], start_row: int,
                 width: int = 720, height: int = 360, row_step: int = 22) -> int:
    """Stack images vertically at column A from start_row. Returns next free row."""
    r = start_row
    for p in paths:
        try:
            img = XLImage(str(p))
            img.width = width; img.height = height
            ws.add_image(img, f"A{r}")
        except Exception:
            pass
        r += row_step
    return r


def apply_car_conditional_formatting(ws: Worksheet, first_row: int, last_row: int,
                                     col_indices: list[int]) -> None:
    if last_row < first_row:
        return
    for ci in col_indices:
        letter = get_column_letter(ci)
        rng = f"{letter}{first_row}:{letter}{last_row}"
        rule = ColorScaleRule(
            start_type="num", start_value=-0.05, start_color="C00000",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="num", end_value=0.05, end_color="00B050",
        )
        ws.conditional_formatting.add(rng, rule)


def write_table(ws: Worksheet, df, cols: list[str], start_row: int) -> tuple[int, int]:
    """Write df[cols] as a header-styled table. Returns (first_data_row, last_data_row)."""
    for c, h in enumerate(cols, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    r = start_row
    first_data = r + 1
    for _, row in df.reindex(columns=cols).iterrows():
        r += 1
        for c, k in enumerate(cols, start=1):
            v = row[k]
            if isinstance(v, float):
                v = round(v, 4)
            ws.cell(row=r, column=c, value=v)
    return first_data, r


def new_workbook(overview_title: str = "Overview") -> tuple[Workbook, Worksheet]:
    wb = Workbook()
    ws = wb.active
    ws.title = overview_title
    return wb, ws
